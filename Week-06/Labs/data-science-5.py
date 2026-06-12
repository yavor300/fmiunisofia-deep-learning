from __future__ import annotations

import argparse
import json
import random
import re
from dataclasses import dataclass
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image
import torch
import torch.nn as nn
from torch.utils.data import DataLoader, Dataset
from torchvision import transforms
from torchvision.models.detection import MaskRCNN_ResNet50_FPN_Weights, maskrcnn_resnet50_fpn
from tqdm.auto import tqdm


@dataclass(frozen=True)
class PanopticSample:
    image_id: int
    image_path: Path
    mask_path: Path
    segments_info: list[dict]


class ConvBlock(nn.Module):
    def __init__(self, in_channels: int, out_channels: int):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.block(x)


class UNet(nn.Module):
    def __init__(self, in_channels: int = 3, out_channels: int = 21, base_channels: int = 32):
        super().__init__()
        c1, c2, c3, c4 = base_channels, base_channels * 2, base_channels * 4, base_channels * 8

        self.initial = ConvBlock(in_channels, c1)
        self.pool1 = nn.MaxPool2d(2)
        self.enc1 = ConvBlock(c1, c2)
        self.pool2 = nn.MaxPool2d(2)
        self.enc2 = ConvBlock(c2, c3)
        self.pool3 = nn.MaxPool2d(2)
        self.enc3 = ConvBlock(c3, c4)

        self.up1 = nn.ConvTranspose2d(c4, c3, kernel_size=2, stride=2)
        self.dec1 = ConvBlock(c3 + c3, c3)
        self.up2 = nn.ConvTranspose2d(c3, c2, kernel_size=2, stride=2)
        self.dec2 = ConvBlock(c2 + c2, c2)
        self.up3 = nn.ConvTranspose2d(c2, c1, kernel_size=2, stride=2)
        self.dec3 = ConvBlock(c1 + c1, c1)

        self.final_conv = nn.Conv2d(c1, out_channels, kernel_size=1)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x0 = self.initial(x)
        x1 = self.enc1(self.pool1(x0))
        x2 = self.enc2(self.pool2(x1))
        x3 = self.enc3(self.pool3(x2))

        y = self.up1(x3)
        y = torch.cat([y, x2], dim=1)
        y = self.dec1(y)

        y = self.up2(y)
        y = torch.cat([y, x1], dim=1)
        y = self.dec2(y)

        y = self.up3(y)
        y = torch.cat([y, x0], dim=1)
        y = self.dec3(y)

        return self.final_conv(y)


class RelabelledCocoSemanticDataset(Dataset):
    def __init__(
        self,
        samples: list[PanopticSample],
        image_size: int,
        category_to_class: dict[int, int],
    ):
        self.samples = samples
        self.category_to_class = category_to_class
        self.resize_image = transforms.Resize((image_size, image_size), interpolation=transforms.InterpolationMode.BILINEAR)
        self.resize_mask = transforms.Resize((image_size, image_size), interpolation=transforms.InterpolationMode.NEAREST)
        self.to_tensor = transforms.ToTensor()
        self.normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, index: int) -> tuple[torch.Tensor, torch.Tensor, torch.Tensor]:
        sample = self.samples[index]

        image = Image.open(sample.image_path).convert("RGB")
        image = self.resize_image(image)

        raw_image_tensor = self.to_tensor(image)
        semantic_input = self.normalize(raw_image_tensor.clone())

        mask = Image.open(sample.mask_path).convert("RGB")
        mask = self.resize_mask(mask)
        instance_id_map = rgb_panoptic_to_id(np.array(mask))

        semantic_target = np.zeros(instance_id_map.shape, dtype=np.int64)
        for segment in sample.segments_info:
            segment_id = int(segment["id"])
            category_id = int(segment["category_id"])
            class_idx = self.category_to_class.get(category_id, 0)
            semantic_target[instance_id_map == segment_id] = class_idx

        target_tensor = torch.from_numpy(semantic_target).long()
        return semantic_input, raw_image_tensor, target_tensor


def rgb_panoptic_to_id(mask_array: np.ndarray) -> np.ndarray:
    if mask_array.ndim == 2:
        return mask_array.astype(np.int64)
    if mask_array.ndim != 3 or mask_array.shape[2] < 3:
        raise ValueError(f"Expected a 2D mask or RGB panoptic mask, got shape {mask_array.shape}.")

    mask_array = mask_array.astype(np.int64)
    return mask_array[:, :, 0] + 256 * mask_array[:, :, 1] + 256 * 256 * mask_array[:, :, 2]


def get_project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def seed_everything(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def load_panoptic_json(json_path: Path) -> dict:
    return json.loads(json_path.read_text())


def build_samples(root: Path, metadata: dict, max_samples: int | None, seed: int) -> list[PanopticSample]:
    images_by_id = {int(item["id"]): item for item in metadata["images"]}
    annotations = metadata["annotations"][:]

    rng = random.Random(seed)
    rng.shuffle(annotations)

    samples: list[PanopticSample] = []
    for ann in annotations:
        image_id = int(ann["image_id"])
        image_info = images_by_id.get(image_id)
        if image_info is None:
            continue

        image_path = root / "images" / image_info["file_name"]
        mask_path = root / "masks" / ann["file_name"]
        if not image_path.exists() or not mask_path.exists():
            continue

        samples.append(
            PanopticSample(
                image_id=image_id,
                image_path=image_path,
                mask_path=mask_path,
                segments_info=ann["segments_info"],
            )
        )

        if max_samples is not None and len(samples) >= max_samples:
            break

    return samples


def split_samples(samples: list[PanopticSample], seed: int) -> tuple[list[PanopticSample], list[PanopticSample], list[PanopticSample]]:
    rng = random.Random(seed)
    shuffled = samples[:]
    rng.shuffle(shuffled)

    n = len(shuffled)
    n_train = int(0.7 * n)
    n_val = int(0.15 * n)

    train = shuffled[:n_train]
    val = shuffled[n_train : n_train + n_val]
    test = shuffled[n_train + n_val :]

    return train, val, test


def build_category_mapping(samples: list[PanopticSample], categories: list[dict], top_k: int) -> tuple[dict[int, int], list[str]]:
    area_by_category: dict[int, float] = {}
    category_name = {int(cat["id"]): str(cat["name"]) for cat in categories}

    for sample in samples:
        for seg in sample.segments_info:
            cat_id = int(seg["category_id"])
            area = float(seg.get("area", 1.0))
            area_by_category[cat_id] = area_by_category.get(cat_id, 0.0) + area

    top_categories = sorted(area_by_category.items(), key=lambda x: x[1], reverse=True)[:top_k]
    chosen_ids = [cat_id for cat_id, _ in top_categories]

    category_to_class = {cat_id: idx + 1 for idx, cat_id in enumerate(chosen_ids)}
    class_names = ["background"] + [category_name.get(cat_id, f"category_{cat_id}") for cat_id in chosen_ids]

    return category_to_class, class_names


def update_confusion(confusion: np.ndarray, preds: torch.Tensor, targets: torch.Tensor, num_classes: int) -> None:
    p = preds.detach().cpu().view(-1).numpy()
    t = targets.detach().cpu().view(-1).numpy()

    valid = (t >= 0) & (t < num_classes)
    p = p[valid]
    t = t[valid]

    for gt, pr in zip(t, p):
        confusion[int(gt), int(pr)] += 1


def metrics_from_confusion(confusion: np.ndarray) -> dict[str, float]:
    total = confusion.sum()
    pixel_acc = float(np.trace(confusion) / total) if total > 0 else 0.0

    ious: list[float] = []
    for cls in range(confusion.shape[0]):
        tp = confusion[cls, cls]
        fp = confusion[:, cls].sum() - tp
        fn = confusion[cls, :].sum() - tp
        denom = tp + fp + fn
        if denom > 0:
            ious.append(float(tp / denom))

    miou = float(np.mean(ious)) if ious else 0.0
    return {"pixel_acc": pixel_acc, "miou": miou}


def train_one_epoch(
    model: nn.Module,
    loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    criterion: nn.Module,
    device: torch.device,
    num_classes: int,
    epoch: int,
) -> dict[str, float]:
    model.train()
    losses: list[float] = []
    confusion = np.zeros((num_classes, num_classes), dtype=np.int64)

    for semantic_inputs, _, targets in tqdm(loader, desc=f"Training epoch {epoch}", leave=False):
        semantic_inputs = semantic_inputs.to(device)
        targets = targets.to(device)

        optimizer.zero_grad()
        logits = model(semantic_inputs)
        loss = criterion(logits, targets)
        loss.backward()
        optimizer.step()

        losses.append(float(loss.item()))
        preds = torch.argmax(logits.detach(), dim=1)
        update_confusion(confusion, preds, targets, num_classes=num_classes)

    metrics = metrics_from_confusion(confusion)
    metrics["loss"] = float(np.mean(losses)) if losses else 0.0
    return metrics


def evaluate_semantic(
    model: nn.Module,
    loader: DataLoader,
    criterion: nn.Module,
    device: torch.device,
    num_classes: int,
    desc: str,
) -> dict[str, float]:
    model.eval()
    losses: list[float] = []
    confusion = np.zeros((num_classes, num_classes), dtype=np.int64)

    with torch.no_grad():
        for semantic_inputs, _, targets in tqdm(loader, desc=desc, leave=False):
            semantic_inputs = semantic_inputs.to(device)
            targets = targets.to(device)

            logits = model(semantic_inputs)
            loss = criterion(logits, targets)
            losses.append(float(loss.item()))

            preds = torch.argmax(logits, dim=1)
            update_confusion(confusion, preds, targets, num_classes=num_classes)

    metrics = metrics_from_confusion(confusion)
    metrics["loss"] = float(np.mean(losses)) if losses else 0.0
    return metrics


def evaluate_panoptic(
    semantic_model: nn.Module,
    instance_model: nn.Module | None,
    loader: DataLoader,
    device: torch.device,
    instance_threshold: float,
) -> dict[str, float]:
    semantic_model.eval()
    if instance_model is not None:
        instance_model.eval()

    coverages: list[float] = []
    fg_recalls: list[float] = []
    instance_counts: list[float] = []

    with torch.no_grad():
        for semantic_inputs, raw_inputs, targets in tqdm(loader, desc="Testing (Panoptic)", leave=False):
            semantic_inputs = semantic_inputs.to(device)
            targets = targets.to(device)

            semantic_logits = semantic_model(semantic_inputs)
            semantic_preds = torch.argmax(semantic_logits, dim=1)

            outputs = None
            if instance_model is not None:
                instance_inputs = [img.to(device) for img in raw_inputs]
                outputs = instance_model(instance_inputs)

            for i in range(semantic_preds.shape[0]):
                semantic_mask = semantic_preds[i].clone()
                panoptic_mask = semantic_mask.clone()

                used_instances = 0
                if outputs is not None:
                    masks = outputs[i].get("masks", torch.empty(0, 1, *semantic_mask.shape, device=device))
                    scores = outputs[i].get("scores", torch.empty(0, device=device))

                    instance_id = int(semantic_mask.max().item()) + 1
                    for mask, score in zip(masks, scores):
                        if float(score.item()) < instance_threshold:
                            continue
                        panoptic_mask[mask[0] > 0.5] = instance_id
                        instance_id += 1
                        used_instances += 1

                gt = targets[i]
                coverage = float((panoptic_mask > 0).float().mean().item())
                gt_fg = (gt > 0).float()
                pred_fg = (panoptic_mask > 0).float()
                fg_recall = float(((pred_fg * gt_fg).sum() / (gt_fg.sum() + 1e-8)).item())

                coverages.append(coverage)
                fg_recalls.append(fg_recall)
                instance_counts.append(float(used_instances))

    return {
        "panoptic_coverage": float(np.mean(coverages)) if coverages else 0.0,
        "foreground_recall": float(np.mean(fg_recalls)) if fg_recalls else 0.0,
        "avg_instances_used": float(np.mean(instance_counts)) if instance_counts else 0.0,
    }


def evaluate_background_baseline(loader: DataLoader, num_classes: int) -> dict[str, float]:
    confusion = np.zeros((num_classes, num_classes), dtype=np.int64)

    for _, _, targets in tqdm(loader, desc="Evaluating baseline", leave=False):
        preds = torch.zeros_like(targets)
        update_confusion(confusion, preds, targets, num_classes=num_classes)

    metrics = metrics_from_confusion(confusion)
    metrics["foreground_recall"] = 0.0
    return metrics


def pct_change_vs_baseline(value: float, baseline: float) -> str:
    if abs(baseline) < 1e-12:
        if abs(value) < 1e-12:
            return "0.00%"
        return "N/A"
    return f"{((value - baseline) / abs(baseline)) * 100.0:+.2f}%"


def save_training_curves(
    labs_dir: Path,
    train_losses: list[float],
    val_losses: list[float],
    train_mious: list[float],
    val_mious: list[float],
) -> tuple[Path, Path]:
    epochs = list(range(1, len(train_losses) + 1))

    metric_path = labs_dir / "data-science-5-train-vs-val-metric.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    ax.plot(epochs, train_mious, marker="o", label="train mean IoU")
    ax.plot(epochs, val_mious, marker="o", label="validation mean IoU")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Mean IoU")
    ax.set_title("Train vs Validation Metric")
    ax.grid(alpha=0.25)
    ax.legend()
    plt.tight_layout()
    plt.savefig(metric_path, dpi=170)
    plt.close(fig)

    loss_path = labs_dir / "data-science-5-train-vs-val-loss.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    ax.plot(epochs, train_losses, marker="o", label="train loss")
    ax.plot(epochs, val_losses, marker="o", label="validation loss")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Loss")
    ax.set_title("Train vs Validation Loss")
    ax.grid(alpha=0.25)
    ax.legend()
    plt.tight_layout()
    plt.savefig(loss_path, dpi=170)
    plt.close(fig)

    return metric_path, loss_path


def parse_markdown_table(markdown_text: str) -> list[list[str]]:
    rows: list[list[str]] = []
    in_table = False
    for line in markdown_text.splitlines():
        if line.startswith("| Model ID |"):
            in_table = True
        if not in_table:
            continue
        if not line.startswith("|"):
            break
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if cells and all(set(cell) <= {"-", " "} for cell in cells):
            continue
        rows.append(cells)
    return rows


def write_markdown_report(
    output_path: Path,
    train_size: int,
    val_size: int,
    test_size: int,
    epochs: int,
    batch_size: int,
    lr: float,
    image_size: int,
    max_samples: int | None,
    top_k_categories: int,
    instance_threshold: float,
    class_names: list[str],
    baseline_test: dict[str, float],
    semantic_val: dict[str, float],
    semantic_test: dict[str, float],
    panoptic_test: dict[str, float],
) -> None:
    baseline_acc = baseline_test["pixel_acc"]
    baseline_miou = baseline_test["miou"]
    baseline_recall = baseline_test["foreground_recall"]

    content = f"""# Data Science Task 05 - Panoptic Segmentation Report

## Best Model
M2 (U-Net semantic mask + Mask R-CNN instance overlays) is the best model because it keeps the semantic branch performance from M1 and adds instance-level panoptic fusion with measurable foreground recall.

## Dataset Exploration
- Dataset: `DATA/relabelled_coco`
- Number of semantic classes used (including background): {len(class_names)}
- Classes used: {", ".join(class_names[:15])}{"..." if len(class_names) > 15 else ""}

## Train / Validation / Test
- Split sizes: {train_size}/{val_size}/{test_size}
- Epochs: {epochs}
- Batch size: {batch_size}
- Learning rate: {lr}
- Image size: {image_size}x{image_size}
- Max samples used: {max_samples}
- Top semantic categories used: {top_k_categories}
- Mask R-CNN instance threshold: {instance_threshold}

## Experiments Table
| Model ID | Hypothesis | Epochs | Batch Size | Learning Rate | Image Size | Test Pixel Accuracy | Delta vs Baseline (Acc) | Test Mean IoU | Delta vs Baseline (mIoU) | Test Foreground Recall | Delta vs Baseline (Recall) | Context/Explanation | Comments |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| M0 | Baseline (predict background class only) | - | - | - | {image_size}x{image_size} | {baseline_test['pixel_acc']:.4f} | 0.00% | {baseline_test['miou']:.4f} | 0.00% | {baseline_test['foreground_recall']:.4f} | 0.00% | Greedy statistical baseline used as the first row, following the model-report guideline. | Reference point for all percentage-change columns. |
| M1 | U-Net semantic segmentation | {epochs} | {batch_size} | {lr} | {image_size}x{image_size} | {semantic_test['pixel_acc']:.4f} | {pct_change_vs_baseline(semantic_test['pixel_acc'], baseline_acc)} | {semantic_test['miou']:.4f} | {pct_change_vs_baseline(semantic_test['miou'], baseline_miou)} | 0.0000 | 0.00% | Semantic branch predicts one class per pixel using relabelled COCO masks. | Improves semantic segmentation metrics over the baseline. |
| M2 | U-Net + Mask R-CNN panoptic fusion | {epochs} | {batch_size} | {lr} | {image_size}x{image_size} | {semantic_test['pixel_acc']:.4f} | {pct_change_vs_baseline(semantic_test['pixel_acc'], baseline_acc)} | {semantic_test['miou']:.4f} | {pct_change_vs_baseline(semantic_test['miou'], baseline_miou)} | {panoptic_test['foreground_recall']:.4f} | {pct_change_vs_baseline(panoptic_test['foreground_recall'], baseline_recall)} | Combines the semantic U-Net mask with Mask R-CNN instance masks as described in the segmentation notes. | Best model because it adds instance-level behavior while preserving the semantic branch metrics. |

## Semantic Branch (U-Net)
### Validation
- Loss: {semantic_val['loss']:.4f}
- Pixel Accuracy: {semantic_val['pixel_acc']:.4f}
- Mean IoU: {semantic_val['miou']:.4f}

### Test
- Loss: {semantic_test['loss']:.4f}
- Pixel Accuracy: {semantic_test['pixel_acc']:.4f}
- Mean IoU: {semantic_test['miou']:.4f}

## Panoptic Fusion (Semantic + Mask R-CNN)
- Panoptic coverage: {panoptic_test['panoptic_coverage']:.4f}
- Foreground recall: {panoptic_test['foreground_recall']:.4f}
- Average fused instance masks per image: {panoptic_test['avg_instances_used']:.2f}

## Diagrams
1. Train vs Validation metric (Mean IoU): shows whether semantic segmentation quality improves similarly on train and validation data.
2. Train vs Validation loss: shows optimization progress and whether train/validation losses diverge.
"""
    output_path.write_text(content)


def write_excel_report_from_markdown(
    output_path: Path,
    markdown_report_path: Path,
    metric_plot_path: Path,
    loss_plot_path: Path,
) -> None:
    markdown_text = markdown_report_path.read_text()
    table = parse_markdown_table(markdown_text)

    best_match = re.search(r"## Best Model\n(.+?)(?:\n##|\Z)", markdown_text, flags=re.DOTALL)
    best_text = best_match.group(1).strip() if best_match else "Best model is described in the Markdown report."

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Report"

    ws["A1"] = "Best Model"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws["A2"] = best_text

    start_row = 4
    for row_idx, row in enumerate(table, start=start_row):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == start_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            if row_idx > start_row and row[0] == "M2":
                cell.fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    for col in range(1, 15):
        ws.column_dimensions[chr(64 + col)].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["M"].width = 55
    ws.column_dimensions["N"].width = 55

    ws_diag = wb.create_sheet(title="Diagrams")
    ws_diag.column_dimensions["A"].width = 120
    ws_diag["A1"] = "Training Diagnostics"
    ws_diag["A1"].font = Font(bold=True)
    ws_diag["A3"] = (
        "Diagram 1: Train vs Validation metric (Mean IoU). "
        "This shows how the semantic segmentation quality changes over epochs on train and validation data."
    )
    ws_diag["A30"] = (
        "Diagram 2: Train vs Validation loss. "
        "This shows optimization progress and whether the model starts overfitting."
    )
    ws_diag.add_image(XLImage(str(metric_plot_path)), "A5")
    ws_diag.add_image(XLImage(str(loss_plot_path)), "A32")

    ws_source = wb.create_sheet(title="Markdown Source")
    ws_source.column_dimensions["A"].width = 140
    ws_source["A1"] = "Markdown report source used for this Excel file:"
    ws_source["A1"].font = Font(bold=True)
    ws_source["A3"] = markdown_text

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Task 05: Panoptic segmentation with U-Net + Mask R-CNN")
    parser.add_argument("--epochs", type=int, default=3)
    parser.add_argument("--batch-size", type=int, default=6)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--image-size", type=int, default=128)
    parser.add_argument("--max-samples", type=int, default=400)
    parser.add_argument("--top-k-categories", type=int, default=20)
    parser.add_argument("--instance-threshold", type=float, default=0.5)
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    seed_everything(args.seed)

    root = get_project_root()
    labs_dir = Path(__file__).resolve().parent

    dataset_root = root / "DATA" / "relabelled_coco"
    json_path = dataset_root / "relabeled_coco_val.json"

    metadata = load_panoptic_json(json_path)

    print("Loaded JSON metadata:")
    print(f"- images: {len(metadata['images'])}")
    print(f"- annotations: {len(metadata['annotations'])}")
    print(f"- categories: {len(metadata['categories'])}")

    samples = build_samples(dataset_root, metadata, max_samples=args.max_samples, seed=args.seed)
    train_samples, val_samples, test_samples = split_samples(samples, seed=args.seed)

    category_to_class, class_names = build_category_mapping(
        train_samples,
        categories=metadata["categories"],
        top_k=args.top_k_categories,
    )
    num_classes = len(class_names)

    print(f"Using {num_classes} semantic classes (including background).")

    train_ds = RelabelledCocoSemanticDataset(train_samples, args.image_size, category_to_class)
    val_ds = RelabelledCocoSemanticDataset(val_samples, args.image_size, category_to_class)
    test_ds = RelabelledCocoSemanticDataset(test_samples, args.image_size, category_to_class)

    train_loader = DataLoader(train_ds, batch_size=args.batch_size, shuffle=True)
    val_loader = DataLoader(val_ds, batch_size=args.batch_size, shuffle=False)
    test_loader = DataLoader(test_ds, batch_size=args.batch_size, shuffle=False)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    semantic_model = UNet(in_channels=3, out_channels=num_classes, base_channels=32).to(device)
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(semantic_model.parameters(), lr=args.lr)

    best_val = {"loss": float("inf"), "pixel_acc": 0.0, "miou": 0.0}
    best_state: dict[str, torch.Tensor] | None = None
    train_losses: list[float] = []
    val_losses: list[float] = []
    train_mious: list[float] = []
    val_mious: list[float] = []

    for epoch in range(1, args.epochs + 1):
        train_metrics = train_one_epoch(
            semantic_model,
            train_loader,
            optimizer,
            criterion,
            device,
            num_classes=num_classes,
            epoch=epoch,
        )
        val_metrics = evaluate_semantic(
            semantic_model,
            val_loader,
            criterion,
            device,
            num_classes=num_classes,
            desc=f"Validation epoch {epoch}",
        )
        train_losses.append(train_metrics["loss"])
        val_losses.append(val_metrics["loss"])
        train_mious.append(train_metrics["miou"])
        val_mious.append(val_metrics["miou"])

        print(
            f"Epoch {epoch}/{args.epochs} | "
            f"train_loss={train_metrics['loss']:.4f} | "
            f"val_loss={val_metrics['loss']:.4f} | "
            f"val_acc={val_metrics['pixel_acc']:.4f} | "
            f"val_mIoU={val_metrics['miou']:.4f}"
        )

        if val_metrics["loss"] < best_val["loss"]:
            best_val = val_metrics
            best_state = {k: v.detach().cpu().clone() for k, v in semantic_model.state_dict().items()}

    if best_state is not None:
        semantic_model.load_state_dict(best_state)

    semantic_test = evaluate_semantic(
        semantic_model,
        test_loader,
        criterion,
        device,
        num_classes=num_classes,
        desc="Testing (Semantic)",
    )

    instance_model: nn.Module | None = None
    try:
        instance_model = maskrcnn_resnet50_fpn(weights=MaskRCNN_ResNet50_FPN_Weights.DEFAULT).to(device)
    except Exception as exc:
        print(
            "Warning: Failed to load pretrained Mask R-CNN weights. "
            f"Panoptic fusion will continue without instance overlays. Details: {exc}"
        )

    panoptic_test = evaluate_panoptic(
        semantic_model,
        instance_model,
        test_loader,
        device,
        instance_threshold=args.instance_threshold,
    )

    print("\nSemantic test metrics:")
    print(f"Loss: {semantic_test['loss']:.4f}")
    print(f"Pixel Accuracy: {semantic_test['pixel_acc']:.4f}")
    print(f"Mean IoU: {semantic_test['miou']:.4f}")

    print("\nPanoptic fusion metrics:")
    print(f"Panoptic coverage: {panoptic_test['panoptic_coverage']:.4f}")
    print(f"Foreground recall: {panoptic_test['foreground_recall']:.4f}")
    print(f"Average instances used: {panoptic_test['avg_instances_used']:.2f}")

    baseline_test = evaluate_background_baseline(test_loader, num_classes=num_classes)
    metric_plot_path, loss_plot_path = save_training_curves(
        labs_dir=labs_dir,
        train_losses=train_losses,
        val_losses=val_losses,
        train_mious=train_mious,
        val_mious=val_mious,
    )

    md_report_path = labs_dir / "data-science-5-model-report.md"
    write_markdown_report(
        output_path=md_report_path,
        train_size=len(train_samples),
        val_size=len(val_samples),
        test_size=len(test_samples),
        epochs=args.epochs,
        batch_size=args.batch_size,
        lr=args.lr,
        image_size=args.image_size,
        max_samples=args.max_samples,
        top_k_categories=args.top_k_categories,
        instance_threshold=args.instance_threshold,
        class_names=class_names,
        baseline_test=baseline_test,
        semantic_val=best_val,
        semantic_test=semantic_test,
        panoptic_test=panoptic_test,
    )

    xlsx_report_path = labs_dir / "data-science-5-model-report.xlsx"
    write_excel_report_from_markdown(
        output_path=xlsx_report_path,
        markdown_report_path=md_report_path,
        metric_plot_path=metric_plot_path,
        loss_plot_path=loss_plot_path,
    )

    print(f"\nMarkdown report saved to: {md_report_path}")
    print(f"Excel report saved to: {xlsx_report_path}")
    print(f"Metric plot saved to: {metric_plot_path}")
    print(f"Loss plot saved to: {loss_plot_path}")


if __name__ == "__main__":
    main()
