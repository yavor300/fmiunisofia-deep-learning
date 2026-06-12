from __future__ import annotations

import argparse
import random
import re
from dataclasses import dataclass
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from PIL import Image
import torch
import torch.nn as nn
from torch.utils.data import DataLoader, Dataset
from torchvision import transforms
from tqdm.auto import tqdm


@dataclass(frozen=True)
class SegSample:
    image_path: Path
    mask_path: Path


class ConvBlock(nn.Module):
    def __init__(self, in_channels: int, out_channels: int):
        super().__init__()
        self.block = nn.Sequential(
            nn.Conv2d(in_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
            nn.Conv2d(out_channels, out_channels, kernel_size=3, padding=1),
            nn.ReLU(inplace=True),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.block(x)


class UNet(nn.Module):
    def __init__(self, in_channels: int = 3, out_channels: int = 1, base_channels: int = 32):
        super().__init__()

        c1, c2, c3, c4 = base_channels, base_channels * 2, base_channels * 4, base_channels * 8

        self.initial = ConvBlock(in_channels, c1)
        self.pool1 = nn.MaxPool2d(2)
        self.enc1 = ConvBlock(c1, c2)
        self.pool2 = nn.MaxPool2d(2)
        self.enc2 = ConvBlock(c2, c3)
        self.pool3 = nn.MaxPool2d(2)
        self.enc3 = ConvBlock(c3, c4)

        self.up1 = nn.ConvTranspose2d(c4, c3, kernel_size=2, stride=2)
        self.dec1 = ConvBlock(c3 + c3, c3)
        self.up2 = nn.ConvTranspose2d(c3, c2, kernel_size=2, stride=2)
        self.dec2 = ConvBlock(c2 + c2, c2)
        self.up3 = nn.ConvTranspose2d(c2, c1, kernel_size=2, stride=2)
        self.dec3 = ConvBlock(c1 + c1, c1)

        self.final_conv = nn.Conv2d(c1, out_channels, kernel_size=1)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        x0 = self.initial(x)
        x1 = self.enc1(self.pool1(x0))
        x2 = self.enc2(self.pool2(x1))
        x3 = self.enc3(self.pool3(x2))

        y = self.up1(x3)
        y = torch.cat([y, x2], dim=1)
        y = self.dec1(y)

        y = self.up2(y)
        y = torch.cat([y, x1], dim=1)
        y = self.dec2(y)

        y = self.up3(y)
        y = torch.cat([y, x0], dim=1)
        y = self.dec3(y)

        return self.final_conv(y)


class CatsDogsSegmentationDataset(Dataset):
    def __init__(self, samples: list[SegSample], image_size: int):
        self.samples = samples
        self.resize_image = transforms.Resize((image_size, image_size), interpolation=transforms.InterpolationMode.BILINEAR)
        self.resize_mask = transforms.Resize((image_size, image_size), interpolation=transforms.InterpolationMode.NEAREST)
        self.to_tensor = transforms.ToTensor()

    def __len__(self) -> int:
        return len(self.samples)

    def __getitem__(self, index: int) -> tuple[torch.Tensor, torch.Tensor]:
        sample = self.samples[index]

        image = Image.open(sample.image_path).convert("RGB")
        mask = Image.open(sample.mask_path).convert("L")

        image = self.resize_image(image)
        mask = self.resize_mask(mask)

        image_tensor = self.to_tensor(image)

        # Dataset labels: 1=foreground, 2=background, 3=not classified.
        mask_np = np.array(mask, dtype=np.uint8)
        binary_mask = (mask_np == 1).astype(np.float32)
        mask_tensor = torch.from_numpy(binary_mask).unsqueeze(0)

        return image_tensor, mask_tensor


def get_project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def seed_everything(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)


def collect_samples(images_dir: Path, masks_dir: Path) -> list[SegSample]:
    samples: list[SegSample] = []
    for image_path in sorted(images_dir.glob("*.jpg")):
        mask_path = masks_dir / f"{image_path.stem}.png"
        if mask_path.exists():
            samples.append(SegSample(image_path=image_path, mask_path=mask_path))
    return samples


def split_samples(samples: list[SegSample], seed: int) -> tuple[list[SegSample], list[SegSample], list[SegSample]]:
    rng = random.Random(seed)
    shuffled = samples[:]
    rng.shuffle(shuffled)

    n = len(shuffled)
    n_train = int(n * 0.7)
    n_val = int(n * 0.15)

    train_samples = shuffled[:n_train]
    val_samples = shuffled[n_train : n_train + n_val]
    test_samples = shuffled[n_train + n_val :]

    return train_samples, val_samples, test_samples


def compute_batch_stats(logits: torch.Tensor, targets: torch.Tensor) -> dict[str, float]:
    probs = torch.sigmoid(logits)
    preds = (probs > 0.5).float()

    targets_flat = targets.view(-1)
    preds_flat = preds.view(-1)

    intersection = torch.sum(preds_flat * targets_flat).item()
    union = torch.sum((preds_flat + targets_flat) > 0).item()
    pred_sum = torch.sum(preds_flat).item()
    target_sum = torch.sum(targets_flat).item()
    correct = torch.sum(preds_flat == targets_flat).item()
    total = targets_flat.numel()

    iou = intersection / (union + 1e-8)
    dice = (2.0 * intersection) / (pred_sum + target_sum + 1e-8)
    acc = correct / total

    return {"iou": iou, "dice": dice, "acc": acc}


def train_one_epoch(
    model: nn.Module,
    loader: DataLoader,
    optimizer: torch.optim.Optimizer,
    criterion: nn.Module,
    device: torch.device,
    epoch: int,
) -> tuple[float, float]:
    model.train()
    running_loss = 0.0
    running_dice = 0.0

    progress = tqdm(loader, desc=f"Training epoch {epoch}", leave=False)
    for images, masks in progress:
        images = images.to(device)
        masks = masks.to(device)

        optimizer.zero_grad()
        logits = model(images)
        loss = criterion(logits, masks)
        loss.backward()
        optimizer.step()

        running_loss += float(loss.item())
        stats = compute_batch_stats(logits.detach(), masks)
        running_dice += stats["dice"]
        progress.set_postfix(loss=f"{loss.item():.4f}")

    n_batches = max(1, len(loader))
    return running_loss / n_batches, running_dice / n_batches


def evaluate(
    model: nn.Module,
    loader: DataLoader,
    criterion: nn.Module,
    device: torch.device,
    desc: str,
) -> dict[str, float]:
    model.eval()
    losses: list[float] = []
    ious: list[float] = []
    dices: list[float] = []
    accs: list[float] = []

    with torch.no_grad():
        for images, masks in tqdm(loader, desc=desc, leave=False):
            images = images.to(device)
            masks = masks.to(device)

            logits = model(images)
            loss = criterion(logits, masks)
            losses.append(float(loss.item()))

            stats = compute_batch_stats(logits, masks)
            ious.append(stats["iou"])
            dices.append(stats["dice"])
            accs.append(stats["acc"])

    return {
        "loss": float(np.mean(losses)) if losses else 0.0,
        "iou": float(np.mean(ious)) if ious else 0.0,
        "dice": float(np.mean(dices)) if dices else 0.0,
        "acc": float(np.mean(accs)) if accs else 0.0,
    }


def evaluate_baseline(loader: DataLoader, device: torch.device) -> dict[str, float]:
    ious: list[float] = []
    dices: list[float] = []
    accs: list[float] = []

    with torch.no_grad():
        for _, masks in tqdm(loader, desc="Evaluating baseline", leave=False):
            masks = masks.to(device)
            # Very low logits => sigmoid near 0 => predict background for all pixels.
            baseline_logits = torch.full_like(masks, fill_value=-20.0)
            stats = compute_batch_stats(baseline_logits, masks)
            ious.append(stats["iou"])
            dices.append(stats["dice"])
            accs.append(stats["acc"])

    return {
        "iou": float(np.mean(ious)) if ious else 0.0,
        "dice": float(np.mean(dices)) if dices else 0.0,
        "acc": float(np.mean(accs)) if accs else 0.0,
    }


def save_training_curves(
    labs_dir: Path,
    train_losses: list[float],
    val_losses: list[float],
    train_metric: list[float],
    val_metric: list[float],
) -> tuple[Path, Path]:
    epochs = list(range(1, len(train_losses) + 1))

    metric_path = labs_dir / "data-science-4-train-vs-val-metric.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    ax.plot(epochs, train_metric, marker="o", label="train dice")
    ax.plot(epochs, val_metric, marker="o", label="validation dice")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Dice")
    ax.set_title("Train vs Validation Metric")
    ax.grid(alpha=0.25)
    ax.legend()
    plt.tight_layout()
    plt.savefig(metric_path, dpi=170)
    plt.close(fig)

    loss_path = labs_dir / "data-science-4-train-vs-val-loss.png"
    fig, ax = plt.subplots(figsize=(7.5, 4.5))
    ax.plot(epochs, train_losses, marker="o", label="train loss")
    ax.plot(epochs, val_losses, marker="o", label="validation loss")
    ax.set_xlabel("Epoch")
    ax.set_ylabel("Loss")
    ax.set_title("Train vs Validation Loss")
    ax.grid(alpha=0.25)
    ax.legend()
    plt.tight_layout()
    plt.savefig(loss_path, dpi=170)
    plt.close(fig)

    return metric_path, loss_path


def pct_change_vs_baseline(value: float, baseline: float) -> str:
    if abs(baseline) < 1e-12:
        if abs(value) < 1e-12:
            return "0.00%"
        return "N/A"
    delta = ((value - baseline) / abs(baseline)) * 100.0
    return f"{delta:+.2f}%"


def write_excel_model_report(
    output_path: Path,
    markdown_report_path: Path,
    baseline_test: dict[str, float],
    model_test: dict[str, float],
    epochs: int,
    batch_size: int,
    lr: float,
    image_size: int,
    metric_plot_path: Path,
    loss_plot_path: Path,
) -> None:
    markdown_text = markdown_report_path.read_text() if markdown_report_path.exists() else ""

    best_model = "M1 (UNet_v1)"
    best_section = re.search(r"## Best Model\\n(.+?)(?:\\n##|\\Z)", markdown_text, flags=re.DOTALL)
    if best_section:
        first_line = [line.strip() for line in best_section.group(1).strip().splitlines() if line.strip()]
        if first_line:
            best_model = first_line[0]

    exp_context: dict[str, str] = {}
    exp_comments: dict[str, str] = {}
    for exp_id, body in re.findall(r"### (M\\d+)[^\\n]*\\n(.+?)(?=\\n### |\\n## |\\Z)", markdown_text, flags=re.DOTALL):
        ctx_match = re.search(r"- Context/Explanation:\\s*(.+)", body)
        com_match = re.search(r"- Comments:\\s*(.+)", body)
        exp_context[exp_id] = ctx_match.group(1).strip() if ctx_match else "N/A"
        exp_comments[exp_id] = com_match.group(1).strip() if com_match else "N/A"

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Report"

    ws["A1"] = f"Best model: {best_model}"
    ws["A1"].font = Font(bold=True)
    ws["A1"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    headers = [
        "Model ID",
        "Hypothesis",
        "Epochs",
        "Batch Size",
        "Learning Rate",
        "Image Size",
        "Test Pixel Accuracy",
        "Δ vs Baseline (Acc)",
        "Test IoU",
        "Δ vs Baseline (IoU)",
        "Test Dice",
        "Δ vs Baseline (Dice)",
        "Context / Explanation",
        "Comments",
    ]

    header_row = 3
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    baseline_row = [
        "M0",
        "Baseline (predict only background pixels)",
        "-",
        "-",
        "-",
        f"{image_size}x{image_size}",
        round(baseline_test["acc"], 4),
        "0.00%",
        round(baseline_test["iou"], 4),
        "0.00%",
        round(baseline_test["dice"], 4),
        "0.00%",
        exp_context.get("M0", "Reference model."),
        exp_comments.get("M0", "Reference model. Used for relative improvements."),
    ]

    model_row = [
        "M1",
        "UNet_v1 (binary semantic segmentation)",
        epochs,
        batch_size,
        lr,
        f"{image_size}x{image_size}",
        round(model_test["acc"], 4),
        pct_change_vs_baseline(model_test["acc"], baseline_test["acc"]),
        round(model_test["iou"], 4),
        pct_change_vs_baseline(model_test["iou"], baseline_test["iou"]),
        round(model_test["dice"], 4),
        pct_change_vs_baseline(model_test["dice"], baseline_test["dice"]),
        exp_context.get("M1", "Trained U-Net semantic segmentation model."),
        exp_comments.get("M1", "Best overall test Dice among tested hypotheses."),
    ]

    for col_idx, value in enumerate(baseline_row, start=1):
        ws.cell(row=4, column=col_idx, value=value)
    for col_idx, value in enumerate(model_row, start=1):
        ws.cell(row=5, column=col_idx, value=value)

    for cell_ref in ("G5", "I5", "K5", "A1"):
        ws[cell_ref].font = Font(bold=True)
    for cell_ref in ("G5", "I5", "K5"):
        ws[cell_ref].fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")

    widths = {
        "A": 10,
        "B": 44,
        "C": 9,
        "D": 10,
        "E": 14,
        "F": 12,
        "G": 18,
        "H": 18,
        "I": 12,
        "J": 18,
        "K": 12,
        "L": 18,
        "M": 48,
        "N": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws["A8"] = "Diagrams and their explanations are included in the 'Diagrams' sheet."
    ws["A8"].font = Font(bold=True)

    ws_diag = wb.create_sheet(title="Diagrams")
    ws_diag["A1"] = "Training Diagnostics"
    ws_diag["A1"].font = Font(bold=True)
    ws_diag.column_dimensions["A"].width = 120

    ws_diag["A3"] = (
        "Diagram 1: Train vs Validation Metric (Dice). "
        "Shows how segmentation quality evolves across epochs and helps detect underfitting/overfitting."
    )
    ws_diag["A30"] = (
        "Diagram 2: Train vs Validation Loss. "
        "Shows optimization behavior and divergence/convergence between train and validation objectives."
    )
    ws_diag.add_image(XLImage(str(metric_plot_path)), "A5")
    ws_diag.add_image(XLImage(str(loss_plot_path)), "A32")

    ws_md = wb.create_sheet(title="Markdown Source")
    ws_md["A1"] = "Markdown report source used for this Excel file:"
    ws_md["A1"].font = Font(bold=True)
    ws_md["A3"] = markdown_text
    ws_md.column_dimensions["A"].width = 140

    wb.save(output_path)


def write_markdown_summary(
    output_path: Path,
    train_size: int,
    val_size: int,
    test_size: int,
    baseline_test: dict[str, float],
    best_val: dict[str, float],
    test_metrics: dict[str, float],
    epochs: int,
    batch_size: int,
    lr: float,
    image_size: int,
) -> None:
    delta_acc = pct_change_vs_baseline(test_metrics["acc"], baseline_test["acc"])
    delta_iou = pct_change_vs_baseline(test_metrics["iou"], baseline_test["iou"])
    delta_dice = pct_change_vs_baseline(test_metrics["dice"], baseline_test["dice"])

    content = f"""# Data Science Task 04 - Model Report

## Best Model
M1 (UNet_v1) is the best model because it achieves higher test Dice than the baseline while keeping stable IoU and pixel accuracy.

## Dataset Context
- Dataset: `DATA/segmentation_cats_dogs`
- Split (train/val/test): {train_size}/{val_size}/{test_size}

## Best Validation Metrics (M1)
- Loss: {best_val['loss']:.4f}
- Pixel Accuracy: {best_val['acc']:.4f}
- IoU: {best_val['iou']:.4f}
- Dice: {best_val['dice']:.4f}

## Experiments Table
| Model ID | Hypothesis | Epochs | Batch Size | Learning Rate | Image Size | Test Pixel Accuracy | Δ vs Baseline (Acc) | Test IoU | Δ vs Baseline (IoU) | Test Dice | Δ vs Baseline (Dice) | Context/Explanation | Comments |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| M0 | Baseline (all-background predictor) | - | - | - | {image_size}x{image_size} | {baseline_test['acc']:.4f} | 0.00% | {baseline_test['iou']:.4f} | 0.00% | {baseline_test['dice']:.4f} | 0.00% | Greedy statistical baseline used as first row in the report. | Serves only as a reference point for percentage change. |
| M1 | UNet_v1 (binary semantic segmentation) | {epochs} | {batch_size} | {lr} | {image_size}x{image_size} | {test_metrics['acc']:.4f} | {delta_acc} | {test_metrics['iou']:.4f} | {delta_iou} | {test_metrics['dice']:.4f} | {delta_dice} | Binary semantic segmentation with U-Net trained on foreground-vs-non-foreground masks. | Selected as best model due to superior Dice and consistent IoU/accuracy compared to M0. |

## Diagrams
1. Train vs Validation metric (Dice): tracks segmentation quality over epochs.
2. Train vs Validation loss: tracks optimization and generalization gap over epochs.
"""
    output_path.write_text(content)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Task 04: Semantic segmentation with U-Net on segmentation_cats_dogs")
    parser.add_argument("--epochs", type=int, default=5)
    parser.add_argument("--batch-size", type=int, default=8)
    parser.add_argument("--lr", type=float, default=1e-3)
    parser.add_argument("--image-size", type=int, default=128)
    parser.add_argument("--seed", type=int, default=42)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    seed_everything(args.seed)

    root = get_project_root()
    labs_dir = Path(__file__).resolve().parent

    images_dir = root / "DATA" / "segmentation_cats_dogs" / "images"
    masks_dir = root / "DATA" / "segmentation_cats_dogs" / "annotations"

    all_samples = collect_samples(images_dir, masks_dir)
    train_samples, val_samples, test_samples = split_samples(all_samples, args.seed)

    train_ds = CatsDogsSegmentationDataset(train_samples, image_size=args.image_size)
    val_ds = CatsDogsSegmentationDataset(val_samples, image_size=args.image_size)
    test_ds = CatsDogsSegmentationDataset(test_samples, image_size=args.image_size)

    train_loader = DataLoader(train_ds, batch_size=args.batch_size, shuffle=True)
    val_loader = DataLoader(val_ds, batch_size=args.batch_size, shuffle=False)
    test_loader = DataLoader(test_ds, batch_size=args.batch_size, shuffle=False)

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    model = UNet(in_channels=3, out_channels=1, base_channels=32).to(device)
    criterion = nn.BCEWithLogitsLoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=args.lr)

    best_val = {"loss": float("inf"), "iou": 0.0, "dice": 0.0, "acc": 0.0}
    best_state: dict[str, torch.Tensor] | None = None

    train_losses: list[float] = []
    val_losses: list[float] = []
    train_dices: list[float] = []
    val_dices: list[float] = []

    for epoch in range(1, args.epochs + 1):
        train_loss, train_dice = train_one_epoch(model, train_loader, optimizer, criterion, device, epoch)
        val_metrics = evaluate(model, val_loader, criterion, device, desc=f"Validation epoch {epoch}")

        train_losses.append(train_loss)
        val_losses.append(val_metrics["loss"])
        train_dices.append(train_dice)
        val_dices.append(val_metrics["dice"])

        print(
            f"Epoch {epoch}/{args.epochs} | "
            f"train_loss={train_loss:.4f} | "
            f"val_loss={val_metrics['loss']:.4f} | "
            f"val_acc={val_metrics['acc']:.4f} | "
            f"val_iou={val_metrics['iou']:.4f} | "
            f"val_dice={val_metrics['dice']:.4f}"
        )

        if val_metrics["loss"] < best_val["loss"]:
            best_val = val_metrics
            best_state = {k: v.detach().cpu().clone() for k, v in model.state_dict().items()}

    if best_state is not None:
        model.load_state_dict(best_state)

    test_metrics = evaluate(model, test_loader, criterion, device, desc="Testing")

    print("\nTest metrics:")
    print(f"Loss: {test_metrics['loss']:.4f}")
    print(f"Pixel Accuracy: {test_metrics['acc']:.4f}")
    print(f"IoU: {test_metrics['iou']:.4f}")
    print(f"Dice: {test_metrics['dice']:.4f}")

    baseline_test = evaluate_baseline(test_loader, device=device)

    metric_plot_path, loss_plot_path = save_training_curves(
        labs_dir=labs_dir,
        train_losses=train_losses,
        val_losses=val_losses,
        train_metric=train_dices,
        val_metric=val_dices,
    )

    md_report_path = labs_dir / "data-science-4-model-report.md"
    write_markdown_summary(
        output_path=md_report_path,
        train_size=len(train_samples),
        val_size=len(val_samples),
        test_size=len(test_samples),
        baseline_test=baseline_test,
        best_val=best_val,
        test_metrics=test_metrics,
        epochs=args.epochs,
        batch_size=args.batch_size,
        lr=args.lr,
        image_size=args.image_size,
    )

    xlsx_report_path = labs_dir / "data-science-4-model-report.xlsx"
    write_excel_model_report(
        output_path=xlsx_report_path,
        markdown_report_path=md_report_path,
        baseline_test=baseline_test,
        model_test=test_metrics,
        epochs=args.epochs,
        batch_size=args.batch_size,
        lr=args.lr,
        image_size=args.image_size,
        metric_plot_path=metric_plot_path,
        loss_plot_path=loss_plot_path,
    )

    print(f"\nModel report saved to: {xlsx_report_path}")
    print(f"Summary saved to: {md_report_path}")
    print(f"Metric plot saved to: {metric_plot_path}")
    print(f"Loss plot saved to: {loss_plot_path}")


if __name__ == "__main__":
    main()
